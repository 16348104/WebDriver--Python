import time

from selenium import webdriver
from openpyxl import *
import xlrd
from public_jxgl import LoginJXGL


class Test_JXGL():
    def __init__(self):
        # self.driver = webdriver.Chrome()
        # self.driver = webdriver.Firefox()
        # self.driver = webdriver.Ie()
        # Mac os
        # self.driver = webdriver.Safari()
        self.driver = webdriver.Firefox(executable_path='/Users/xdx/PycharmProjects/WebDriver--Python/geckodriver')
        # self.driver = webdriver.Chrome(executable_path='/Users/xdx/PycharmProjects/WebDriver--Python/chromedriver')
        # info模拟环境
        self.driver.get('http://info.syx.thcic.cn')
        self.driver.maximize_window()
        print('测试浏览器:' + self.driver.name)
        time.sleep(3)

        # 登录

    def login(self):
        wb = load_workbook(r'sample.xlsx', read_only=True)
        my_sheet = wb.worksheets[0]
        username = my_sheet.cell(row=2, column=1).value
        password = my_sheet.cell(row=2, column=2).value
        LoginJXGL().userlogin(self.driver, username, password)

    # 填写未评估课程问卷2017012040
    def questionaire_wp(self):
        wb = load_workbook(r'sample.xlsx', read_only=True)
        my_sheet = wb.worksheets[0]
        # A2 = my_sheet["A2"]
        # B2 = my_sheet["B2"]
        # 用cell 函数
        username = my_sheet.cell(row=3, column=1).value
        password = my_sheet.cell(row=3, column=2).value
        LoginJXGL().userlogin(self.driver, username, password)
        LoginJXGL().fill_questionaire_wp(self.driver)
        LoginJXGL().closed(self.driver)

        # 填写已评估课程问卷2017013478

    def questionaire_yp(self):
        wb = load_workbook(r'sample.xlsx', read_only=True)
        my_sheet = wb.worksheets[0]
        # A2 = my_sheet["A2"]
        # B2 = my_sheet["B2"]
        # 用cell 函数
        username = my_sheet.cell(row=2, column=1).value
        password = my_sheet.cell(row=2, column=2).value
        LoginJXGL().userlogin(self.driver, username, password)
        LoginJXGL().fill_questionaire_yp(self.driver)
        LoginJXGL().closed(self.driver)

    # 转换评估课程
    def change(self):
        readbook = xlrd.open_workbook(r'sample.xlsx')
        # writebook = xlwt.Workbook()#打开一个excel
        # sheet = writebook.add_sheet('test')#在打开的excel中添加一个sheet
        # 获取读入的文件的第一个sheet
        table = readbook.sheets()[0]
        username = table.cell(2, 0).value  # 获取2行1列的表格值
        password = table.cell(2, 1).value  # 获取2行2列的表格值
        # print(username, password)
        LoginJXGL().userlogin(self.driver, username, password)
        LoginJXGL().change_list(self.driver)
        LoginJXGL().closed(self.driver)

    # 查阅历史问卷
    def evaluation(self):
        wb = load_workbook(r'sample.xlsx', read_only=True)
        my_sheet = wb.worksheets[0]
        # A2 = my_sheet["A2"]
        # B2 = my_sheet["B2"]
        # 用cell 函数
        username = my_sheet.cell(row=2, column=1).value
        password = my_sheet.cell(row=2, column=2).value
        LoginJXGL().userlogin(self.driver, username, password)
        LoginJXGL().view_evaluation(self.driver)
        LoginJXGL().closed(self.driver)
        LoginJXGL().email()


# 执行测试
# Test_JXGL().login()
# Test_JXGL().questionaire_wp()
# Test_JXGL().questionaire_yp()
# Test_JXGL().change()
Test_JXGL().evaluation()
