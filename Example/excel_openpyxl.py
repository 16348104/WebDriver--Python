import datetime
from openpyxl import *

# 有关Excel 2007版之前的excel文件后缀为.xls，最大支持65535行数据，xlrd和xlwt主要应用于07版之前的Excel文件。xlrd可以用来读取.xls和.xlsx文件， xlwt写文件只支持.xls，所以对数据的大小有限制
# 2007版之后的excel文件后缀为.xlsx, 最大支持1048576行，使用openpyxl来弥补xlwt的缺陷来处理大文件。
# 有个比较简单的解决办法就是在数字和日期的单元格内容前加上一个英文的逗号即可。如果数据比较多，也可以批量加英文逗号的前缀

# 只读模式打开文件
wb = load_workbook(r'E:\sample.xlsx', read_only=True)
# 加载存在的 excel 文件: 默认可读写
# wb = load_workbook(r'E:\sample.xlsx')
my_sheet = wb.worksheets[0]
A1 = my_sheet["A2"]
B1 = my_sheet["B2"]
# 用cell 函数
a1_cell = my_sheet.cell(row=2, column=1)
b1_cell = my_sheet.cell(row=2, column=2)
# print(A1.value, B1.value)
print(a1_cell.value, b1_cell.value)

# # grab the active worksheet
# ws = wb.active
#
# # Data can be assigned directly to cells
# ws['A1'] = 42
#
# # Rows can also be appended
# ws.append([1, 2, 3])
#
# # Python types will automatically be converted
# ws['A2'] = datetime.datetime.now()
#
# # Save the file
# wb.save("sample.xlsx")
